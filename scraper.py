import os
import time
import tempfile
import logging
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DEMANDER_URL = "https://sistema.demander.com.br"
LOGIN_EMAIL = os.environ.get("DEMANDER_EMAIL", "")
LOGIN_SENHA = os.environ.get("DEMANDER_SENHA", "")


class DemandScraper:
    """Acessa o Demander via Playwright e extrai dados de clientes."""

    def __init__(self):
        self.email = LOGIN_EMAIL
        self.senha = LOGIN_SENHA
        if not self.email or not self.senha:
            raise ValueError("DEMANDER_EMAIL e DEMANDER_SENHA precisam estar definidos no .env")

    # ------------------------------------------------------------------
    # LOGIN
    # ------------------------------------------------------------------
    def _fazer_login(self, page):
        """Realiza login no Demander."""
        logger.info("Fazendo login no Demander...")
        page.goto(DEMANDER_URL, wait_until="networkidle")

        # Preenche e-mail
        page.fill('input[type="email"], input[name="email"], input[placeholder*="mail"]', self.email)
        # Preenche senha
        page.fill('input[type="password"]', self.senha)
        # Clica em entrar
        page.click('button[type="submit"], button:has-text("Entrar")')

        # Aguarda redirecionamento pós-login
        page.wait_for_url(lambda url: "login" not in url, timeout=15000)
        logger.info("Login realizado com sucesso.")

    # ------------------------------------------------------------------
    # BUSCAR CIDADES DISPONÍVEIS PARA UM ESTADO
    # ------------------------------------------------------------------
    def buscar_cidades(self, estado: str) -> list[str]:
        """
        Navega até o filtro de clientes no Demander,
        seleciona o estado e retorna as cidades disponíveis.
        """
        cidades = []

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            try:
                self._fazer_login(page)
                self._navegar_para_clientes(page)

                # Seleciona o estado no filtro
                self._selecionar_estado(page, estado)

                # Aguarda o select de cidades ser populado
                page.wait_for_timeout(1500)

                # Captura as opções do select de cidade
                cidades = page.eval_on_selector_all(
                    # Ajuste o seletor conforme o HTML real do Demander
                    'select[name*="cidade"], select[id*="cidade"], select[name*="city"]',
                    """elements => {
                        const sel = elements[0];
                        if (!sel) return [];
                        return Array.from(sel.options)
                            .map(o => o.text.trim())
                            .filter(t => t && t !== '' && t !== 'Selecione' && t !== 'Todos');
                    }"""
                )

                logger.info(f"Cidades encontradas em {estado}: {cidades}")

            except PlaywrightTimeout as e:
                logger.error(f"Timeout ao buscar cidades: {e}")
                raise
            finally:
                browser.close()

        return cidades

    # ------------------------------------------------------------------
    # EXPORTAR CLIENTES DE UMA CIDADE
    # ------------------------------------------------------------------
    def exportar_clientes(self, estado: str, cidade: str) -> str:
        """
        Filtra clientes por estado + cidade e retorna
        o caminho de um arquivo .xlsx temporário.
        """
        clientes = []

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            try:
                self._fazer_login(page)
                self._navegar_para_clientes(page)

                # Aplica filtros
                self._selecionar_estado(page, estado)
                page.wait_for_timeout(1000)
                self._selecionar_cidade(page, cidade)

                # Clica em pesquisar/filtrar
                page.click('button:has-text("Pesquisar"), button:has-text("Filtrar"), button:has-text("Buscar")')
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(1000)

                # Coleta todos os clientes (paginação incluída)
                clientes = self._coletar_todas_paginas(page)
                logger.info(f"Total de clientes coletados: {len(clientes)}")

            except PlaywrightTimeout as e:
                logger.error(f"Timeout ao exportar clientes: {e}")
                raise
            finally:
                browser.close()

        return self._gerar_excel(clientes, estado, cidade)

    # ------------------------------------------------------------------
    # NAVEGAÇÃO INTERNA
    # ------------------------------------------------------------------
    def _navegar_para_clientes(self, page):
        """
        Navega até a tela de listagem de clientes.
        Ajuste a URL/seletor conforme o menu real do Demander.
        """
        # Tenta navegar direto pela URL — ajuste o path se necessário
        try:
            page.goto(f"{DEMANDER_URL}/clientes", wait_until="networkidle", timeout=10000)
            return
        except Exception:
            pass

        # Fallback: clica no menu
        try:
            page.click('a:has-text("Clientes"), nav a:has-text("Cliente")')
            page.wait_for_load_state("networkidle")
        except Exception as e:
            logger.warning(f"Não encontrou menu de clientes: {e}")

    def _selecionar_estado(self, page, estado: str):
        """Seleciona o estado no filtro."""
        try:
            # Tenta pelo select padrão
            page.select_option(
                'select[name*="estado"], select[id*="estado"], select[name*="state"], select[name*="uf"]',
                label=estado
            )
        except Exception:
            try:
                page.select_option(
                    'select[name*="estado"], select[id*="estado"], select[name*="state"], select[name*="uf"]',
                    value=estado
                )
            except Exception as e:
                logger.warning(f"Não conseguiu selecionar estado via select: {e}")

    def _selecionar_cidade(self, page, cidade: str):
        """Seleciona a cidade no filtro."""
        try:
            page.select_option(
                'select[name*="cidade"], select[id*="cidade"], select[name*="city"]',
                label=cidade
            )
        except Exception:
            try:
                page.select_option(
                    'select[name*="cidade"], select[id*="cidade"], select[name*="city"]',
                    value=cidade
                )
            except Exception as e:
                logger.warning(f"Não conseguiu selecionar cidade via select: {e}")

    # ------------------------------------------------------------------
    # COLETA DE DADOS (COM PAGINAÇÃO)
    # ------------------------------------------------------------------
    def _coletar_todas_paginas(self, page) -> list[dict]:
        """Percorre todas as páginas da tabela e coleta os dados."""
        todos = []
        pagina = 1

        while True:
            logger.info(f"Coletando página {pagina}...")
            dados = self._coletar_pagina_atual(page)
            todos.extend(dados)

            # Tenta ir para próxima página
            proximo = page.query_selector(
                'a:has-text("Próximo"), button:has-text("Próximo"), '
                'a[rel="next"], .pagination .next:not(.disabled)'
            )
            if not proximo or not proximo.is_enabled():
                break

            proximo.click()
            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(800)
            pagina += 1

        return todos

    def _coletar_pagina_atual(self, page) -> list[dict]:
        """Extrai linhas da tabela de clientes da página atual."""
        clientes = []

        try:
            # Pega cabeçalhos
            headers = page.eval_on_selector_all(
                "table thead th, table thead td",
                "els => els.map(e => e.innerText.trim())"
            )

            # Pega linhas
            linhas = page.query_selector_all("table tbody tr")

            for linha in linhas:
                colunas = linha.query_selector_all("td")
                if not colunas:
                    continue
                valores = [col.inner_text().strip() for col in colunas]
                if headers:
                    cliente = {headers[i]: valores[i] for i in range(min(len(headers), len(valores)))}
                else:
                    cliente = {f"Coluna {i+1}": v for i, v in enumerate(valores)}
                clientes.append(cliente)

        except Exception as e:
            logger.error(f"Erro ao coletar página: {e}")

        return clientes

    # ------------------------------------------------------------------
    # GERAÇÃO DO EXCEL
    # ------------------------------------------------------------------
    def _gerar_excel(self, clientes: list[dict], estado: str, cidade: str) -> str:
        """Gera arquivo .xlsx formatado e retorna o caminho."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{cidade} - {estado}"

        # Estilo do cabeçalho
        header_fill = PatternFill("solid", fgColor="1D9E75")  # verde Demander
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center")

        if not clientes:
            ws.append(["Nenhum cliente encontrado para esta cidade."])
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            wb.save(tmp.name)
            return tmp.name

        # Cabeçalho
        colunas = list(clientes[0].keys())
        for col_idx, col_name in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Dados
        for row_idx, cliente in enumerate(clientes, start=2):
            for col_idx, col_name in enumerate(colunas, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cliente.get(col_name, ""))
                cell.alignment = Alignment(vertical="center")
                # Linhas alternadas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F1F5F9")

        # Ajuste de largura das colunas
        for col_idx, col_name in enumerate(colunas, start=1):
            max_len = max(
                len(str(col_name)),
                *[len(str(c.get(col_name, ""))) for c in clientes]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        ws.row_dimensions[1].height = 22

        # Freeze cabeçalho
        ws.freeze_panes = "A2"

        # Salva em arquivo temporário
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        logger.info(f"Excel salvo em: {tmp.name}")
        return tmp.name
